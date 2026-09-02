#!/usr/bin/env python3
"""Генерация таблицы тренировок для импорта в Google Таблицы."""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Колонка A — упражнения; далее 4 недели × 5 полей без разрывов
EX_COL = 1
WEEKS = 4
FIELDS_PER_WEEK = 5

YELLOW = PatternFill("solid", fgColor="FFF2CC")
LIGHT_BLUE = PatternFill("solid", fgColor="D9EAF7")
LIGHT_GREY = PatternFill("solid", fgColor="F3F3F3")

THIN = Side(style="thin", color="000000")
THICK = Side(style="thick", color="000000")

thin_border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
week_outer = Border(left=THICK, right=THICK, top=THIN, bottom=THIN)

SUB_HEADERS = ["Подх", "П-ры\nСек", "Вес", "Отд\nых", "RPE"]
WEEK_NAMES = ["НЕДЕЛЯ 1", "НЕДЕЛЯ 2", "НЕДЕЛЯ 3", "НЕДЕЛЯ 4"]

HEADER_ROW = 1
SUBHEADER_ROW = 2
DATA_START_ROW = 3

WEEK_COLS = []
for w in range(WEEKS):
    start = EX_COL + 1 + w * FIELDS_PER_WEEK
    WEEK_COLS.append((start, start + FIELDS_PER_WEEK - 1))

GRIP_DAY = [
    ("day", "Вторник"),
    (
        "ex",
        "Висы активный хват 20мм",
        [("3", "10с", "", "1-2м", ""), ("3", "7с", "", "1-2м", ""), ("3", "на MAX", "", "1-2м", "")],
    ),
    ("ex", "Подтягивания на турнике", [("", "", "", "", "")]),
    ("ex", "Щипок блок всей ладонью", [("", "", "", "", "")]),
    ("ex", "Брусья", [("", "", "", "", "")]),
    ("ex", "Щипок блок только 1-й фалангой", [("", "", "", "", "")]),
    ("ex", "Т-разводки TRX/Кольца", [("", "", "", "", "")]),
    ("ex", "Y-разводки TRX/Кольца", [("", "", "", "", "")]),
    ("ex", "Наружная ротация", [("", "", "", "", "")]),
    ("acwr", None),
]

STRENGTH_DAY = [
    ("day", "Четверг"),
    ("ex", "Подтягивания мощно", [("3", "3", "100%", "1-2м", "")]),
    ("mun", "МУН", "90мин"),
    ("category", "Интервалы"),
    ("ex", "Присед с ящика", [("", "", "", "", "")]),
    ("ex", "Планка", [("", "", "", "", "")]),
    ("ex", "Боковая планка адДукция", [("", "", "", "", "")]),
    ("ex", "Боковая планка абДукция", [("", "", "", "", "")]),
    ("acwr", None),
]

SATURDAY = [("day", "Суббота"), *GRIP_DAY[1:]]
SUNDAY = [("day", "Воскресенье"), *STRENGTH_DAY[1:]]


def set_cell(ws, row, col, value="", fill=None, font=None, align=None, border=None, merge_to_col=None):
    cell = ws.cell(row=row, column=col, value=value)
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    if align:
        cell.alignment = align
    if border:
        cell.border = border
    if merge_to_col and merge_to_col > col:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_to_col)
        for c in range(col, merge_to_col + 1):
            mc = ws.cell(row=row, column=c)
            if fill:
                mc.fill = fill
            if border:
                mc.border = border
            if align:
                mc.alignment = align
    return cell


def write_headers(ws):
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    bold = Font(bold=True)

    set_cell(ws, HEADER_ROW, EX_COL, "НЕДЕЛЯ", YELLOW, bold, center, thin_border)
    for name, (sc, ec) in zip(WEEK_NAMES, WEEK_COLS):
        set_cell(ws, HEADER_ROW, sc, name, YELLOW, bold, center, week_outer, merge_to_col=ec)

    set_cell(ws, SUBHEADER_ROW, EX_COL, "Упражнение", LIGHT_BLUE, bold, center, thin_border)
    for sc, _ec in WEEK_COLS:
        for i, hdr in enumerate(SUB_HEADERS):
            set_cell(ws, SUBHEADER_ROW, sc + i, hdr, YELLOW, bold, center, thin_border)

    ws.row_dimensions[HEADER_ROW].height = 22
    ws.row_dimensions[SUBHEADER_ROW].height = 36


def write_acwr(ws, start_row):
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    bold = Font(bold=True)

    set_cell(ws, start_row, EX_COL, "ACWR", LIGHT_GREY, bold, left, thin_border)
    for sc, ec in WEEK_COLS:
        set_cell(ws, start_row, sc, "", LIGHT_GREY, border=thin_border, merge_to_col=ec)

    set_cell(
        ws,
        start_row + 1,
        EX_COL,
        "Субъективная оценка тяжести тренировки",
        None,
        None,
        left,
        thin_border,
    )
    for sc, ec in WEEK_COLS:
        for c in range(sc, ec + 1):
            set_cell(ws, start_row + 1, c, "", None, border=thin_border)

    set_cell(
        ws,
        start_row + 2,
        EX_COL,
        "Время тренировки в минутах",
        None,
        None,
        left,
        thin_border,
    )
    for sc, ec in WEEK_COLS:
        for c in range(sc, ec + 1):
            set_cell(ws, start_row + 2, c, "", None, border=thin_border)

    return start_row + 3


def write_day_block(ws, start_row, day_data):
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    bold = Font(bold=True)
    link_font = Font(color="0563C1", underline="single")

    row = start_row
    for item in day_data:
        kind = item[0]

        if kind == "day":
            title = item[1]
            set_cell(ws, row, EX_COL, title, YELLOW, bold, center, thin_border)
            for sc, ec in WEEK_COLS:
                set_cell(ws, row, sc, title, YELLOW, bold, center, week_outer, merge_to_col=ec)
            ws.row_dimensions[row].height = 24
            row += 1

        elif kind == "category":
            set_cell(ws, row, EX_COL, item[1], LIGHT_GREY, bold, left, thin_border)
            for sc, ec in WEEK_COLS:
                set_cell(ws, row, sc, "", LIGHT_GREY, border=thin_border, merge_to_col=ec)
            row += 1

        elif kind == "mun":
            set_cell(ws, row, EX_COL, item[1], None, bold, left, thin_border)
            for sc, ec in WEEK_COLS:
                set_cell(ws, row, sc, item[2], None, None, center, thin_border, merge_to_col=ec)
            row += 1

        elif kind == "ex":
            name, sets_data = item[1], item[2]
            set_cell(ws, row, EX_COL, name, None, link_font, left, thin_border)
            if sets_data:
                data = sets_data[0]
                for week_idx, (sc, _ec) in enumerate(WEEK_COLS):
                    vals = data if week_idx == 0 else ("", "", "", "", "")
                    for i, val in enumerate(vals):
                        set_cell(ws, row, sc + i, val, None, None, center, thin_border)
            row += 1
            for extra in sets_data[1:]:
                set_cell(ws, row, EX_COL, "", None, None, left, thin_border)
                for week_idx, (sc, _ec) in enumerate(WEEK_COLS):
                    vals = extra if week_idx == 0 else ("", "", "", "", "")
                    for i, val in enumerate(vals):
                        set_cell(ws, row, sc + i, val, None, None, center, thin_border)
                row += 1

        elif kind == "acwr":
            row = write_acwr(ws, row)
            row += 1

    return row


def setup_columns(ws):
    ws.column_dimensions[get_column_letter(EX_COL)].width = 34
    for sc, ec in WEEK_COLS:
        for c in range(sc, ec + 1):
            ws.column_dimensions[get_column_letter(c)].width = 9
    ws.freeze_panes = f"{get_column_letter(EX_COL + 1)}{DATA_START_ROW}"


def add_rpe_validation(ws, max_row):
    dv = DataValidation(
        type="list",
        formula1='"1,2,3,4,5,6,7,8,9,10"',
        allow_blank=True,
        showDropDown=False,
    )
    dv.error = "Выберите значение RPE от 1 до 10"
    dv.errorTitle = "Неверное значение"
    ws.add_data_validation(dv)
    for sc, ec in WEEK_COLS:
        rpe_col = get_column_letter(ec)
        dv.add(f"{rpe_col}{DATA_START_ROW}:{rpe_col}{max_row}")


# --- Лист «Инструкции» ---
ACCENT_BLUE = PatternFill("solid", fgColor="1A73E8")
ACCENT_LIGHT = PatternFill("solid", fgColor="E8F0FE")
SECTION_FILL = PatternFill("solid", fgColor="F8F9FA")
RPE_FILLS = {
    "10": PatternFill("solid", fgColor="FCE8E6"),
    "9": PatternFill("solid", fgColor="FEF0E6"),
    "8": PatternFill("solid", fgColor="FFF8E1"),
    "7": PatternFill("solid", fgColor="E8F5E9"),
}

INSTRUCTION_SECTIONS = [
    (
        "1. Разминка.",
        'а. Перед лазанием желательно использовать полный блок с листа "разминка" но он имеет '
        "рекомендательный характер, упражнения можно заменять. НО висы и разминку пальцев делать!\n"
        "б. Пред силовыми разминку производим локально, весь блок делать не надо. Т.е. Если делаем "
        "упражнения на верх, то делаем немного общей разминки по верху тела, а потом упражнение которое "
        "собираемся делать в облегченном формате. Например: планируем подтягивания 30кг*5. разминка: "
        "общая на верх тела(чуть мобильности, работа по плечам и ротаторам плеча 1-2 подхода, без утомления) "
        "затем подтягивания без веса 5-8 раз, с 10-15 кг 5 раз, с 20-25 кг 3р и только потом 30кг*5 раз. "
        "По той же схеме пальцы и ноги.",
        90,
    ),
    (
        "2. RPE-",
        "Это субъективная оценка количества повторов/секунд в запасе после выполнения подхода.",
        28,
    ),
    (
        "4. Обозначения A1,A2. B1,B2.",
        "Это значит, что упражнения с одинаковой буквой выполняются по очереди, а цифра обозначает какое "
        "упражнение первое, а какое 2-е/3-е. Упражнения обозначенные одинаковой буквой и цифрой( A1,A1) "
        "выполняются так: выполнить все подходы упражнения, что расположено выше в таблице, затем то, что ниже.",
        55,
    ),
    (
        "5. Отдых в упражнениях с обозначением A1,A2 и т.д.",
        "Интервал отдыха A1,A2 2мин. Это означает что суммарное время отдыха на 2 упражнения 2 минуты. "
        "После первого упражнения отдых 1 минута. Затем, второе упражнение и еще минута отдыха.",
        45,
    ),
]

RPE_ROWS = [
    ("10", "нет запаса вообще"),
    ("9", "1 повтор или 2-3 сек в запасе."),
    ("8", "2 повтора или 4-6 сек в запасе."),
    ("7", "3 возможно 4 или 7-10сек и т.д."),
]

RPE_EXAMPLE = (
    "Пример. Подтянулся 5 раз с 20 кг,все ровно и красиво, в конце оцениваешь сколько еще повторов "
    "смог бы сделать. И, предположим, 2 повтора еще можно выполнить, значит RPE 8."
)


def write_instructions_sheet(wb):
    ws = wb.create_sheet("Инструкции")
    text_col_end = 6  # A-F как в оригинале

    title_font = Font(name="Arial", bold=True, size=16, color="FFFFFF")
    section_font = Font(name="Arial", bold=True, size=11, color="174EA6")
    body_font = Font(name="Arial", size=11, color="202124")
    wrap_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    wrap_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    section_align = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Заголовок
    set_cell(ws, 1, 1, "Инструкции", ACCENT_BLUE, title_font, wrap_center, thin_border, merge_to_col=text_col_end)
    ws.row_dimensions[1].height = 36

    row = 2
    for title, body, height in INSTRUCTION_SECTIONS:
        set_cell(ws, row, 1, title, SECTION_FILL, section_font, section_align, thin_border)
        set_cell(ws, row, 2, body, None, body_font, wrap_left, thin_border, merge_to_col=text_col_end)
        ws.row_dimensions[row].height = height
        row += 1

        if title == "2. RPE-":
            # Таблица RPE
            set_cell(ws, row, 2, "RPE", ACCENT_LIGHT, Font(name="Arial", bold=True, size=10), wrap_center, thin_border)
            set_cell(ws, row, 3, "Описание", ACCENT_LIGHT, Font(name="Arial", bold=True, size=10), wrap_center, thin_border, merge_to_col=text_col_end)
            ws.row_dimensions[row].height = 22
            row += 1
            for rpe, desc in RPE_ROWS:
                set_cell(ws, row, 2, rpe, RPE_FILLS[rpe], Font(name="Arial", bold=True, size=11), wrap_center, thin_border)
                set_cell(ws, row, 3, desc, PatternFill("solid", fgColor="FAFAFA"), body_font, wrap_left, thin_border, merge_to_col=text_col_end)
                ws.row_dimensions[row].height = 22
                row += 1
            set_cell(ws, row, 2, RPE_EXAMPLE, None, body_font, wrap_left, thin_border, merge_to_col=text_col_end)
            ws.row_dimensions[row].height = 40
            row += 1

    ws.column_dimensions["A"].width = 22
    for c in range(2, text_col_end + 1):
        ws.column_dimensions[get_column_letter(c)].width = 18
    ws.freeze_panes = "A2"


def add_placeholder_sheets(wb):
    for title in ["Разминка", "Стретчинг 1"]:
        ws = wb.create_sheet(title)
        ws["A1"] = f"Лист «{title}» — заполните по необходимости"
        ws["A1"].font = Font(italic=True, color="666666")


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "Неделя 1-4"

    write_headers(ws)
    row = DATA_START_ROW
    for day_block in [GRIP_DAY, STRENGTH_DAY, SATURDAY, SUNDAY]:
        row = write_day_block(ws, row, day_block)
        row += 1

    setup_columns(ws)
    add_rpe_validation(ws, row + 5)
    add_placeholder_sheets(wb)
    write_instructions_sheet(wb)

    out = "/Users/maksimizrailev/Documents/hexlet_studying/ski-guide-Italy/Тренировочный_журнал_4_недели.xlsx"
    wb.save(out)

    import shutil
    dest = "/Users/maksimizrailev/Documents/training-files/workout_journal_4weeks.xlsx"
    shutil.copy(out, dest)
    print(out)
    print(dest)


if __name__ == "__main__":
    main()
